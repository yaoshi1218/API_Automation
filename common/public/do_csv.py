import csv
from openpyxl import load_workbook, workbook
from common.public.my_log import MyLog
file_name = '测试.csv'
rows=[]
fields=('a','b','c')
for i in range(100000):
    rows.append(('1','2','3'))
def write_csv(file_name,fields,rows):
    # 写入csv文件
    with open(file_name,'w',newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        # 写入字段
        csvwriter.writerow(fields)
        # 写入行数据
        csvwriter.writerows(rows)
        print('成功')
#
# log = MyLog()  # 实例化MyLog类
# wb = load_workbook('积分上传_20191224153451.xlsx')
# sheet = wb['会员积分列表']
# for i in range(100000):
#     sheet.cell(i+2, 1).value = '018S0Z6PXV5'  # 将值写入表格
#     sheet.cell(i + 2, 2).value = '1'  # 将值写入表格
#     sheet.cell(i + 2, 3).value = '2019-12-28'  # 将值写入表格
#     sheet.cell(i + 2, 4).value = '吃腻了房' # 将值写入表格
#     log.info('第{}行成功写入'.format(i+1))
# wb.save('积分上传_20191224153451.xlsx')  # 保存文件
# wb.close()  # 关闭文件


# if __name__ == "__main__":
#     write_csv(file_name,fields,rows)



