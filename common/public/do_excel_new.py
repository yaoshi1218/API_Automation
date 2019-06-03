# -*- coding: utf-8 -*-
# @Time    : 2019/3/21 11:01
# @File    : do_excel_new.py
from openpyxl import load_workbook,workbook
from my_log import MyLog
from my_config import ReadConfig
from project_path import *
log=MyLog()#实例化MyLog类

class DoExcel:
    def __init__(self,file_name):
        '''
        这是一个读取表格数据，将数据写入表格、新建表格的类
        file_name 表格名
        sheet_name 表单名'''
        self.file_name = file_name


    def get_tel(self):
        '''这是一个读取手机号码的函数'''
        try:
            wb=load_workbook(self.file_name)
            sheet=wb['tel']
            tel=sheet.cell(2,1).value
            wb.close()
        except Exception as e :
            log.error('读取手机号错误{}'.format(e))
            raise e
        return tel

    def update_tel(self,new_tel):
        '''这是一个写回手机号码的函数'''
        try:
            wb=load_workbook(self.file_name)
            sheet=wb['tel']
            sheet.cell(2,1).value=new_tel
            wb.save(self.file_name)
            log.info('手机号码保存成功')
            wb.close()
        except Exception as e :
            log.error('手机号码写回错误'.format(e))
            raise e


    def read_excel(self,):
        '''读取数据的函数'''
        config = ReadConfig(conf_new_path)  # 配置文件类实例
        module_case = config.read_other('CASE','module_case')#读取配置文件，选择要读取的模块
        # module_case={'recharge':'all','register':[1,2]}

        wb = load_workbook(self.file_name)

        tel = self.get_tel()  # 读取手机号
        final_data=[]#空列表 存储最终的测试用例数据
        for module in module_case:#取出模块的名称
            module_data=[]#每个模块的测试数据
            sheet=wb[module]#定位表单
            for i in range(2,sheet.max_row+1):
                try:
                    row_data = {}#创建存储用例数据的类
                    row_data['case_id'] = sheet.cell(row=i, column=1).value  # 读取单元格的值
                    row_data['module'] = sheet.cell(row=i, column=2).value
                    row_data['title'] = sheet.cell(row=i, column=3).value
                    row_data['url'] = sheet.cell(row=i, column=4).value
                    row_data['method'] = sheet.cell(row=i, column=5).value
                    if sheet.cell(i,6).value.find('tel')!=-1:#这是在查找有没有tel这个字段=-1表示没有
                        row_data['params']=sheet.cell(i,6).value.replace('tel',str(tel))#将tel进行替换成手机号码
                        self.update_tel(tel + 1)
                    else:#如果没有这个tel，就不需要替换
                        row_data['params']=sheet.cell(row=i, column=6).value
                    row_data['expected_result'] = sheet.cell(row=i, column=7).value
                    row_data['sheet_name'] = module#把表单名存起来，写回数据用
                    module_data.append(row_data)#将表格加入到列表中
                    wb.close()  # 关闭文件
                    log.info('正在读取id={}'.format(row_data['case_id']))
                except Exception as e :
                    log.error('读取文件错误'.format(e))
                    raise e

            if module_case[module] == 'all':#判断配置文件中的值，进行相应的取值
                final_data.extend(module_data)#追加进总列表
            else:
                for i in module_case[module]:#循环配置文件的读取的列表
                    final_data.append(module_data[i-1])#根据列表索引取出对应得值加入到final_case
        return final_data
    def write_excel(self,sheet_name,row,column,value):
        '''写入数值的函数
        sheet_name 表单名
        row 写入表格的行数
        column 写入表格的列
        value 写入的值'''
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        try :
            sheet.cell(row,column).value=value#将值写入表格
            wb.save(self.file_name)#保存文件
            wb.close()#关闭文件
            log.info('成功写入')
        except Exception as e :
            log.error('写入文件错误'.format(e))
            raise e
class AddExcel:
    '''创建新的表格文件'''
    def add_excel(self,file_name,sheet_name):
        '''新建表格函数
        file_name 文件名
        sheet_name 表单名'''
        try :
            wb=workbook.Workbook()#新建表格
            wb.create_sheet(sheet_name)#新建表单
            wb.save(file_name)
            log.info('新建成功')
        except Exception as e :
            log.error('新建失败'.format(e))
            raise e
if __name__=="__main__":
    do=DoExcel(test_case_path)
    a=do.read_excel()
    print(a)
    # do.write_excel('login',2,7,'haa')
    # AddExcel().add_excel('test.xlsx','login')